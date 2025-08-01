import pandas as pd
import os
import shutil
from docxtpl import DocxTemplate
from datetime import datetime
from pathlib import Path
import openpyxl             # Para trabajar e iterar en las hojas de los archivos Excel 
import xlwings              # Controla directamente Excel local
import warnings


# -------------------- PARAMETROS --------------------

# Ruta salida
OUTPUT_PATH = './Outputs'

# Ruta del archivo Excel de entrada
EXCEL_PATH = './Parameters/Formato Nombres.xlsx'

# Diccionario de plantillas Word (.docx)
WORD_PATHS = {
    'APP_': './Inputs/Templates/APP_01_XXXXXXXX_XXXXXXXX.docx',
    'AEPP_': './Inputs/Templates/AEPP_01_XXXXXXXX_XXXXXXXX.docx',
}

# Diccionario de plantillas Excel (.xlsx)
EXCEL_TEMPLATES = {
    'EDLLO_': './Inputs/Templates/EDLLO_01_XXXXXXXX_XXXXXXXX_TCS.xlsx',
    'EPP_': './Inputs/Templates/EPP_01_XXXXXXXX_XXXXXXXX_TCS.xlsx',
}

# Fecha actual
FECHA = datetime.today().strftime("%d/%m/%Y")

# No muestra tipo de advertencias.
warnings.simplefilter("ignore", UserWarning)


# -------------------- ELIMINAR Y CREAR CARPETA OUTPUTS --------------------

def eliminar_y_crear_carpeta(path):
    if os.path.exists(path):
        shutil.rmtree(path)
    os.makedirs(path, exist_ok=True)


# ---------------------- CREAR SUB CARPETAS ----------------------

def crear_sub_carpetas(df_datos, output_path, carpetas=None):
    if carpetas is None:
        carpetas = ["01_Diagnostico", "02_Solucion", "03_Pruebas", "04_Instalacion"]

    mensajes = []

    for _, fila in df_datos.iterrows():
        v_oc = fila['OC']
        v_nombre = fila['NOMBRE']
        nombre_carpeta = f"OC_{v_oc}_{v_nombre}"
        ruta_final = os.path.join(output_path, nombre_carpeta)

        try:
            for carpeta in carpetas:
                os.makedirs(os.path.join(ruta_final, carpeta), exist_ok=True)
            print(f"\U00002714 Carpetas creadas en: {ruta_final}")
        except Exception as e:
            mensajes.append(f"\U0000274C Error creando carpetas en {ruta_final}: {e}")

    return mensajes

# -------------------- CREAR DOCUMENTO --------------------

def crear_documentos(df_datos):
    # Unir todos los tipos disponibles
    todos_los_tipos = {**WORD_PATHS, **EXCEL_TEMPLATES}

    # Filtrar solo tipos válidos
    df_datos = df_datos[df_datos['TIPO'].isin(todos_los_tipos.keys())]

    # Obtener lista de nombres únicos
    nombres_oc = df_datos['NOMBRE_OC'].unique()

    for nombre_oc in nombres_oc:
        datos_filtrados = df_datos[df_datos['NOMBRE_OC'] == nombre_oc]

        for _, fila in datos_filtrados.iterrows():
            tipo = fila['TIPO']
            plantilla = todos_los_tipos.get(tipo)

            if not plantilla or not os.path.exists(plantilla):
                print(f"\U0000274C Plantilla no encontrada: {plantilla}")
                continue

            if plantilla.endswith('.docx'):
                doc = DocxTemplate(plantilla)

                contexto = {
                    'Numero_OC': fila['OC'],
                    'Nombre_OC': fila['NOMBRE'],
                    'Desarrollador': fila['DESARROLLADOR'],
                    'Rol': fila['ROL'],
                    'Aplicacion': fila['APLICACION'],
                    'Firma_ET': fila['FIRMA_ET'],
                    'Fecha': FECHA,
                }

                doc.render(contexto)
                output_path = os.path.join(OUTPUT_PATH,f"{nombre_oc}.docx")
                doc.save(output_path)
                print(f"\U00002714 Documento Word generado: {output_path}")

            elif plantilla.endswith('.xlsx'):
                ##wb = xlwings.Book(plantilla)
                wb = openpyxl.load_workbook(plantilla)

                # Modificar EDLLO en la hoja 'Portada'
                if 'Portada' in wb.sheetnames:
                    portada = wb['Portada']
                    portada['F10'] = FECHA
                    portada['F11'] = FECHA
                    portada['F14'] = fila['NOMBRE']
                    portada['F15'] = fila['OC']
                    portada['F23'] = fila['DESARROLLADOR']
                    portada['F21'] = f"{fila['DESARROLLADOR']}/{fila['ROL']}"

                # Modificar EDLLO en la hoja 'Caso 1'
                if 'Caso 1' in wb.sheetnames:
                    caso1 = wb['Caso 1']
                    caso1['C1'] = fila['APLICACION']

                # Modificar EPP en la hoja (1-Est. y Planeación)
                if '1-Est. y Planeación' in wb.sheetnames:
                    Est_y_Plan = wb['1-Est. y Planeación']
                    Est_y_Plan['M3'] = fila['OC']
                    Est_y_Plan['L11'] = fila['QA']
                    Est_y_Plan['Y3'] = fila['NOMBRE']
                    Est_y_Plan['AR3'] = fila['APLICACION']
                    Est_y_Plan['J7'] = FECHA

                # Modificar EPP en la hoja (2-Diseño de Casos Prueba)
                if '2-Diseño de Casos Prueba' in wb.sheetnames:
                    Est_y_Plan = wb['2-Diseño de Casos Prueba']
                    Est_y_Plan['B6'] = fila['APLICACION']


                #Nombrar y guardar documento de salida
                output_path = os.path.join(OUTPUT_PATH, f"{nombre_oc}.xlsx")
                wb.save(output_path)
                print(f"\U00002714 Documento Excel generado: {output_path}")



# ---------------------- MAIN ----------------------

def main():
    print("\U0001F680 Automatización de documentos iniciada \U0001F4DC")

    eliminar_y_crear_carpeta(OUTPUT_PATH)

    try:
        df = pd.read_excel(EXCEL_PATH, sheet_name='Datos')
    except Exception as e:
        print(f"\U0000274C Error al leer el Excel: {e}")
        return

    mensajes = crear_sub_carpetas(df, OUTPUT_PATH)

    crear_documentos(df)

    print("\U00002705 Proceso finalizado correctamente.")


# -------------------- EJECUCIÓN --------------------
if __name__ == '__main__':
    main()
